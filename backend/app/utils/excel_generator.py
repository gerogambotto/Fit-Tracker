from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from app.models.models import Rutina
import re

def generate_rutina_excel(rutina: Rutina):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rutina"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    day_font = Font(bold=True, color="000000")
    day_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Información de la rutina
    ws['A1'] = "Rutina:"
    ws['B1'] = rutina.nombre
    ws['A2'] = "Alumno:"
    ws['B2'] = rutina.alumno.nombre
    ws['A3'] = "Fecha de inicio:"
    ws['B3'] = rutina.fecha_inicio.strftime('%d/%m/%Y') if rutina.fecha_inicio else 'No especificada'
    ws['A4'] = "Entrenamientos/semana:"
    ws['B4'] = rutina.entrenamientos_semana
    ws['A5'] = "Notas:"
    ws['B5'] = rutina.notas or 'Sin notas'
    
    # Organizar ejercicios por día
    dias_ejercicios = {}
    for i in range(1, rutina.entrenamientos_semana + 1):
        dias_ejercicios[i] = []
    
    for ejercicio in rutina.ejercicios:
        dia_match = re.search(r'Día (\d+):', ejercicio.notas or '')
        dia = int(dia_match.group(1)) if dia_match else 1
        if dia in dias_ejercicios:
            dias_ejercicios[dia].append(ejercicio)
    
    row = 7
    
    # Generar contenido por día
    for dia in range(1, rutina.entrenamientos_semana + 1):
        if dias_ejercicios[dia]:
            # Header del día
            ws[f'A{row}'] = f"DÍA {dia}"
            ws[f'A{row}'].font = day_font
            ws[f'A{row}'].fill = day_fill
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].fill = day_fill
            row += 1
            
            # Headers de ejercicios
            ws[f'A{row}'] = "Ejercicio"
            ws[f'B{row}'] = "Series"
            ws[f'C{row}'] = "Repeticiones"
            ws[f'D{row}'] = "Peso (kg)"
            ws[f'E{row}'] = "Descanso (seg)"
            ws[f'F{row}'] = "Notas"
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}', f'E{row}', f'F{row}']:
                ws[cell].font = header_font
                ws[cell].fill = header_fill
            row += 1
            
            # Ejercicios del día
            for ejercicio in dias_ejercicios[dia]:
                ws[f'A{row}'] = ejercicio.ejercicio_base.nombre
                ws[f'B{row}'] = ejercicio.series
                ws[f'C{row}'] = ejercicio.repeticiones
                ws[f'D{row}'] = ejercicio.peso or '-'
                ws[f'E{row}'] = ejercicio.descanso
                notas_limpias = re.sub(r'Día \d+:\s*', '', ejercicio.notas or '')
                ws[f'F{row}'] = notas_limpias or '-'
                row += 1
            
            row += 1  # Espacio entre días
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()