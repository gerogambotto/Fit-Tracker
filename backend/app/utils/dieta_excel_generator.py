from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from app.models.models import Dieta

def generate_dieta_excel(dieta: Dieta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dieta"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
    day_font = Font(bold=True, color="000000")
    day_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    meal_font = Font(bold=True, color="4B0082")
    
    # Información de la dieta
    ws['A1'] = "Dieta:"
    ws['B1'] = dieta.nombre
    ws['A2'] = "Alumno:"
    ws['B2'] = dieta.alumno.nombre if dieta.alumno else 'Sin asignar'
    ws['A3'] = "Fecha de inicio:"
    ws['B3'] = dieta.fecha_inicio.strftime('%d/%m/%Y') if dieta.fecha_inicio else 'No especificada'
    ws['A4'] = "Notas:"
    ws['B4'] = dieta.notas or 'Sin notas'
    
    # Organizar comidas por día
    dias_comidas = {}
    for comida in dieta.comidas:
        dia = comida.dia
        if dia not in dias_comidas:
            dias_comidas[dia] = []
        dias_comidas[dia].append(comida)
    
    row = 6
    
    # Generar contenido por día
    for dia in sorted(dias_comidas.keys()):
        if dias_comidas[dia]:
            # Header del día
            ws[f'A{row}'] = f"MENÚ {dia}"
            ws[f'A{row}'].font = day_font
            ws[f'A{row}'].fill = day_fill
            for col in ['B', 'C', 'D']:
                ws[f'{col}{row}'].fill = day_fill
            row += 1
            
            # Headers de comidas
            ws[f'A{row}'] = "Comida"
            ws[f'B{row}'] = "Alimento"
            ws[f'C{row}'] = "Cantidad (g)"
            ws[f'D{row}'] = "Calorías"
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell].font = header_font
                ws[cell].fill = header_fill
            row += 1
            
            # Comidas del día
            for comida in sorted(dias_comidas[dia], key=lambda x: x.orden):
                first_alimento = True
                for ca in comida.alimentos:
                    if first_alimento:
                        ws[f'A{row}'] = comida.nombre
                        ws[f'A{row}'].font = meal_font
                        first_alimento = False
                    else:
                        ws[f'A{row}'] = ""
                    
                    ws[f'B{row}'] = ca.alimento.nombre
                    ws[f'C{row}'] = ca.cantidad_gramos
                    calorias = (ca.alimento.calorias_100g * ca.cantidad_gramos) / 100
                    ws[f'D{row}'] = f"{calorias:.1f}"
                    row += 1
                
                if not comida.alimentos:
                    ws[f'A{row}'] = comida.nombre
                    ws[f'A{row}'].font = meal_font
                    ws[f'B{row}'] = "Sin alimentos"
                    row += 1
            
            row += 1  # Espacio entre días
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()